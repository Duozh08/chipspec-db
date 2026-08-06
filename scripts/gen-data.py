"""Generate TypeScript data files from Excel."""
import openpyxl
import json
import re
import os

OUT_DIR = os.path.dirname(os.path.abspath(__file__))

# ============================================================
# 1. Generate mobile-chips.ts from 芯片.xlsx
# ============================================================

def parse_tdp(val):
    if not val or val == '—':
        return None
    m = re.search(r'(\d+)', str(val))
    return int(m.group(1)) if m else None

def parse_dims_mm(val):
    """Parse '15.0 × 12.7 mm' -> (15.0, 12.7)"""
    if not val or val == '—':
        return None, None
    m = re.findall(r'(\d+\.?\d*)', str(val).replace('×', 'x'))
    if len(m) >= 2:
        return float(m[0]), float(m[1])
    return None, None

def parse_package_dims(val):
    """Parse '45×42.5mm' -> (45, 42.5)"""
    if not val or val == '—':
        return None, None
    m = re.findall(r'(\d+\.?\d*)', str(val).replace('×', 'x'))
    if len(m) >= 2:
        return float(m[0]), float(m[1])
    return None, None

def parse_transistors(val):
    if not val or val == '—':
        return None
    s = str(val)
    m = re.search(r'(\d+\.?\d*)亿', s)
    if m:
        return int(float(m.group(1)) * 100)  # 亿 -> 百万
    m = re.search(r'(\d+\.?\d*)B', s)
    if m:
        return int(float(m.group(1)) * 1000)
    m = re.search(r'(\d+\.?\d*)M', s)
    if m:
        return int(float(m.group(1)))
    return None

def parse_release(val):
    if not val or val == '—':
        return None
    s = str(val).strip()
    # "2021 Q2" -> "2021-Q2"
    s = re.sub(r'\s+', '-', s)
    return s

def parse_die_area(val):
    if not val or val == '—':
        return None
    m = re.search(r'(\d+\.?\d*)', str(val))
    return float(m.group(1)) if m else None

def parse_dies(die_count, die_dims_str, die_area_str, process):
    """Parse die information from Excel row."""
    dies = []
    if not die_dims_str or die_dims_str == '—':
        # No die dimensions, use area
        area = parse_die_area(die_area_str)
        for i in range(die_count):
            dies.append({
                'name': f'Die {i+1}' if die_count > 1 else 'Monolithic Die',
                'role': 'compute',
                'process': process,
                'areaMm2': area / die_count if area else None,
                'lengthMm': None,
                'widthMm': None,
                'transistorsMillions': None,
            })
        return dies

    dims_str = str(die_dims_str)
    # Check for multi-die format: "CCD: 10.6 × 6.7 mm ×2; IOD: 11.6 × 10.5 mm"
    if ';' in dims_str or 'CCD' in dims_str:
        parts = dims_str.split(';')
        idx = 0
        for part in parts:
            part = part.strip()
            # Parse name
            name_match = re.match(r'(\w+):\s*(.*)', part)
            name = name_match.group(1) if name_match else f'Die {idx+1}'
            rest = name_match.group(2) if name_match else part
            # Check for count "×2"
            count_match = re.search(r'[×x](\d+)', rest)
            count = int(count_match.group(1)) if count_match else 1
            # Parse dims
            m = re.findall(r'(\d+\.?\d*)', rest.split('×')[0] if '×' in rest else rest)
            l, w = None, None
            if len(m) >= 2:
                l, w = float(m[0]), float(m[1])
            area = parse_die_area(die_area_str) if idx == 0 else None

            for c in range(count):
                role = 'compute' if 'CCD' in name or 'Compute' in name else 'io' if 'IOD' in name or 'IO' in name else 'other'
                dies.append({
                    'name': name,
                    'role': role,
                    'process': process,
                    'areaMm2': None,
                    'lengthMm': l,
                    'widthMm': w,
                    'transistorsMillions': None,
                })
            idx += 1
        return dies

    # Single die: "15.0 × 12.7 mm"
    l, w = parse_dims_mm(dims_str)
    area = parse_die_area(die_area_str)
    dies.append({
        'name': 'Monolithic Die',
        'role': 'compute',
        'process': process,
        'areaMm2': area,
        'lengthMm': l,
        'widthMm': w,
        'transistorsMillions': None,
    })
    return dies

def slugify(model):
    s = model.lower()
    s = re.sub(r'[^a-z0-9]+', '-', s)
    s = s.strip('-')
    return s

def brand_key(brand_str):
    return 'intel' if 'Intel' in brand_str else 'amd' if 'AMD' in brand_str else 'nvidia'

def gen_mobile_chips():
    wb = openpyxl.load_workbook('D:/Desktop/芯片.xlsx', data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    chips = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[0] is None:
            continue
        d = dict(zip(headers, row))

        brand_str = str(d.get('品牌', ''))
        brand = brand_key(brand_str)
        model = str(d.get('型号', '')).strip()
        codename = str(d.get('核心代号', '')).strip()
        generation = str(d.get('代际', '')).strip()
        process = str(d.get('制程工艺', '')).strip()
        release = parse_release(d.get('发布时间'))

        pkg_type = str(d.get('封装型号', '')).strip()
        pkg_l, pkg_w = parse_package_dims(d.get('封装尺寸（长×宽）'))

        die_count = int(d.get('Die 数量', 1)) if d.get('Die 数量') else 1
        die_dims = d.get('Die 长×宽')
        die_area = d.get('Die 总面积')

        dies = parse_dies(die_count, die_dims, die_area, process)

        tdp = parse_tdp(d.get('TDP（标准功耗）'))
        max_power = parse_tdp(d.get('最高功耗'))
        load_temp = str(d.get('满载温度', '')).strip()
        notes = str(d.get('备注', '')).strip()
        transistors = parse_transistors(d.get('晶体管总数'))

        chip_id = slugify(model)

        # Build notes with max power
        full_notes = notes
        if max_power:
            full_notes = f'{notes}；最高功耗 {max_power}W' if notes else f'最高功耗 {max_power}W'

        chip = {
            'id': chip_id,
            'brand': brand,
            'category': 'cpu',
            'formFactor': 'mobile',
            'model': model,
            'codename': codename,
            'generation': generation,
            'process': process,
            'release': release,
            'package': {
                'type': pkg_type if pkg_type and pkg_type != '—' else 'Unknown',
                'style': 'bga',
                'lengthMm': pkg_l,
                'widthMm': pkg_w,
            },
            'dies': dies,
            'transistorsMillions': transistors,
            'tdp': tdp,
            'loadTempRange': load_temp if load_temp and load_temp != '—' else None,
            'notes': full_notes if full_notes and full_notes != '—' else None,
            'dataQuality': 'official',
            'sources': [],
        }
        chips.append(chip)

    # Generate TypeScript
    lines = []
    lines.append("import type { Chip, Source } from '../types';\n")
    lines.append("const TPU: Source = { label: 'TechPowerUp GPU Database', url: 'https://www.techpowerup.com/gpu-specs/' };")
    lines.append("const GPU_PKG_MOBILE = { type: 'FCBGA', style: 'bga' as const, lengthMm: null, widthMm: null };\n")
    lines.append("// Auto-generated from 芯片.xlsx — 移动端处理器数据")
    lines.append("// 数据来源：Intel ARK / AMD 官方规格 / WikiChip 等公开来源\n")
    lines.append("export const mobileChips: Chip[] = [")

    for c in chips:
        lines.append("  {")
        lines.append(f"    id: '{c['id']}',")
        lines.append(f"    brand: '{c['brand']}', category: 'cpu', formFactor: 'mobile',")
        lines.append(f"    model: '{c['model']}',")
        lines.append(f"    codename: '{c['codename']}',")
        lines.append(f"    generation: '{c['generation']}',")
        lines.append(f"    process: '{c['process']}',")
        release_str = c['release'] if c['release'] else 'null'
        lines.append(f"    release: {('`' + release_str + '`') if release_str != 'null' else 'null'},")
        lines.append(f"    package: {{ type: '{c['package']['type']}', style: '{c['package']['style']}', lengthMm: {c['package']['lengthMm'] or 'null'}, widthMm: {c['package']['widthMm'] or 'null'} }},")

        # Dies
        dies_ts = []
        for d in c['dies']:
            parts = []
            parts.append(f"name: '{d['name']}'")
            parts.append(f"role: '{d['role']}'")
            parts.append(f"process: {'null' if not d['process'] else chr(39) + d['process'] + chr(39)}")
            parts.append(f"areaMm2: {d['areaMm2'] or 'null'}")
            parts.append(f"lengthMm: {d['lengthMm'] or 'null'}")
            parts.append(f"widthMm: {d['widthMm'] or 'null'}")
            parts.append(f"transistorsMillions: {d['transistorsMillions'] or 'null'}")
            dies_ts.append("{ " + ", ".join(parts) + " }")

        lines.append(f"    dies: [{', '.join(dies_ts)}],")
        lines.append(f"    transistorsMillions: {c['transistorsMillions'] or 'null'},")
        lines.append(f"    tdp: {c['tdp'] or 'null'},")
        temp_str = f"'{c['loadTempRange']}'" if c['loadTempRange'] else 'null'
        lines.append(f"    loadTempRange: {temp_str},")
        notes_str = f"'{c['notes']}'" if c['notes'] else 'null'
        lines.append(f"    notes: {notes_str},")
        lines.append(f"    dataQuality: 'official',")
        lines.append(f"    sources: [],")
        lines.append("  },")

    # 追加手动收录的移动端 GPU（Excel 无 GPU 数据，从旧代码库保留）
    lines.append(MOBILE_GPU_TS)

    lines.append("];")
    return '\n'.join(lines)


MOBILE_GPU_TS = '''
  // ========== 移动端 GPU（手动收录，Excel 未覆盖） ==========
  // --- RTX 50 Laptop (Blackwell) ---
  {
    id: 'nvidia-geforce-rtx-5090-laptop',
    brand: 'nvidia', category: 'gpu', formFactor: 'mobile',
    model: 'GeForce RTX 5090 Laptop',
    codename: 'GB203 (Blackwell)',
    generation: 'RTX 50 Laptop',
    process: 'TSMC 4N (5nm 级)',
    release: '2025-03',
    package: { ...GPU_PKG_MOBILE },
    dies: [{ name: 'GB203 Monolithic Die', role: 'graphics', process: 'TSMC 4N', areaMm2: 378.6, lengthMm: null, widthMm: null, transistorsMillions: 45600, note: '与桌面 RTX 5080 同款 GB203 Die' }],
    transistorsMillions: 45600,
    tdp: 175,
    loadTempRange: '65-80°C',
    notes: 'RTX 50 系旗舰移动 GPU，24GB GDDR7，满血功耗 175W。',
    dataQuality: 'estimated',
    sources: [TPU],
  },
  {
    id: 'nvidia-geforce-rtx-5080-laptop',
    brand: 'nvidia', category: 'gpu', formFactor: 'mobile',
    model: 'GeForce RTX 5080 Laptop',
    codename: 'GB203 (Blackwell)',
    generation: 'RTX 50 Laptop',
    process: 'TSMC 4N (5nm 级)',
    release: '2025-03',
    package: { ...GPU_PKG_MOBILE },
    dies: [{ name: 'GB203 Monolithic Die', role: 'graphics', process: 'TSMC 4N', areaMm2: 378.6, lengthMm: null, widthMm: null, transistorsMillions: 45600, note: '与桌面 RTX 5080 同款 GB203 Die' }],
    transistorsMillions: 45600,
    tdp: 145,
    loadTempRange: '65-80°C',
    notes: 'RTX 5080 Laptop，16GB GDDR7，满血功耗 175W。',
    dataQuality: 'estimated',
    sources: [TPU],
  },
  {
    id: 'nvidia-geforce-rtx-5070-ti-laptop',
    brand: 'nvidia', category: 'gpu', formFactor: 'mobile',
    model: 'GeForce RTX 5070 Ti Laptop',
    codename: 'GB205 (Blackwell)',
    generation: 'RTX 50 Laptop',
    process: 'TSMC 4N (5nm 级)',
    release: '2025-03',
    package: { ...GPU_PKG_MOBILE },
    dies: [{ name: 'GB205 Monolithic Die', role: 'graphics', process: 'TSMC 4N', areaMm2: null, lengthMm: null, widthMm: null, transistorsMillions: null }],
    transistorsMillions: null,
    tdp: 140,
    loadTempRange: '65-80°C',
    notes: 'RTX 5070 Ti Laptop，12GB GDDR7，满血功耗 140W。',
    dataQuality: 'estimated',
    sources: [TPU],
  },
  {
    id: 'nvidia-geforce-rtx-5070-laptop',
    brand: 'nvidia', category: 'gpu', formFactor: 'mobile',
    model: 'GeForce RTX 5070 Laptop',
    codename: 'GB206 (Blackwell)',
    generation: 'RTX 50 Laptop',
    process: 'TSMC 4N (5nm 级)',
    release: '2025-03',
    package: { ...GPU_PKG_MOBILE },
    dies: [{ name: 'GB206 Monolithic Die', role: 'graphics', process: 'TSMC 4N', areaMm2: null, lengthMm: null, widthMm: null, transistorsMillions: null }],
    transistorsMillions: null,
    tdp: 115,
    loadTempRange: '65-80°C',
    notes: 'RTX 5070 Laptop，8GB GDDR7，满血功耗 140W。',
    dataQuality: 'estimated',
    sources: [TPU],
  },
  {
    id: 'nvidia-geforce-rtx-5060-laptop',
    brand: 'nvidia', category: 'gpu', formFactor: 'mobile',
    model: 'GeForce RTX 5060 Laptop',
    codename: 'GB206 (Blackwell)',
    generation: 'RTX 50 Laptop',
    process: 'TSMC 4N (5nm 级)',
    release: '2025-03',
    package: { ...GPU_PKG_MOBILE },
    dies: [{ name: 'GB206 Monolithic Die', role: 'graphics', process: 'TSMC 4N', areaMm2: null, lengthMm: null, widthMm: null, transistorsMillions: null }],
    transistorsMillions: null,
    tdp: 100,
    loadTempRange: '60-78°C',
    notes: 'RTX 5060 Laptop，8GB GDDR7，满血功耗 115W。',
    dataQuality: 'estimated',
    sources: [TPU],
  },
  {
    id: 'nvidia-geforce-rtx-5050-laptop',
    brand: 'nvidia', category: 'gpu', formFactor: 'mobile',
    model: 'GeForce RTX 5050 Laptop',
    codename: 'GB207 (Blackwell)',
    generation: 'RTX 50 Laptop',
    process: 'TSMC 4N (5nm 级)',
    release: '2025-03',
    package: { ...GPU_PKG_MOBILE },
    dies: [{ name: 'GB207 Monolithic Die', role: 'graphics', process: 'TSMC 4N', areaMm2: null, lengthMm: null, widthMm: null, transistorsMillions: null }],
    transistorsMillions: null,
    tdp: 100,
    loadTempRange: '60-78°C',
    notes: 'RTX 5050 Laptop，8GB GDDR7，满血功耗 115W。',
    dataQuality: 'estimated',
    sources: [TPU],
  },
  // --- RTX 40 Laptop (Ada) ---
  {
    id: 'nvidia-geforce-rtx-4090-laptop',
    brand: 'nvidia', category: 'gpu', formFactor: 'mobile',
    model: 'GeForce RTX 4090 Laptop',
    codename: 'AD103 (Ada Lovelace)',
    generation: 'RTX 40 Laptop',
    process: 'TSMC 4N (5nm 级)',
    release: '2023-02',
    package: { ...GPU_PKG_MOBILE },
    dies: [{ name: 'AD103 Monolithic Die', role: 'graphics', process: 'TSMC 4N', areaMm2: 378.6, lengthMm: null, widthMm: null, transistorsMillions: 45900, note: '与桌面 RTX 4080 同款 AD103 Die，TDP 限制后大幅降频' }],
    transistorsMillions: 45900,
    tdp: 150,
    loadTempRange: '72-85°C',
    notes: '移动版 RTX 4090 实为桌面 RTX 4080 同款 AD103 Die，TDP 150W 上下（满血 175W）。',
    dataQuality: 'official',
    sources: [TPU],
  },
  {
    id: 'nvidia-geforce-rtx-4080-laptop',
    brand: 'nvidia', category: 'gpu', formFactor: 'mobile',
    model: 'GeForce RTX 4080 Laptop',
    codename: 'AD104 (Ada Lovelace)',
    generation: 'RTX 40 Laptop',
    process: 'TSMC 4N (5nm 级)',
    release: '2023-02',
    package: { ...GPU_PKG_MOBILE },
    dies: [{ name: 'AD104 Monolithic Die', role: 'graphics', process: 'TSMC 4N', areaMm2: 294.5, lengthMm: null, widthMm: null, transistorsMillions: 35800, note: '与桌面 RTX 4070 Ti 同款 AD104 Die' }],
    transistorsMillions: 35800,
    tdp: 150,
    loadTempRange: '70-83°C',
    notes: '移动版 RTX 4080 使用 AD104 Die，12GB GDDR6X，满血功耗 175W。',
    dataQuality: 'official',
    sources: [TPU],
  },
  {
    id: 'nvidia-geforce-rtx-4070-laptop',
    brand: 'nvidia', category: 'gpu', formFactor: 'mobile',
    model: 'GeForce RTX 4070 Laptop',
    codename: 'AD106 (Ada Lovelace)',
    generation: 'RTX 40 Laptop',
    process: 'TSMC 4N (5nm 级)',
    release: '2023-02',
    package: { ...GPU_PKG_MOBILE },
    dies: [{ name: 'AD106 Monolithic Die', role: 'graphics', process: 'TSMC 4N', areaMm2: 187.8, lengthMm: null, widthMm: null, transistorsMillions: 22900, note: '与桌面 RTX 4060 Ti 同款 AD106 Die' }],
    transistorsMillions: 22900,
    tdp: 115,
    loadTempRange: '68-80°C',
    notes: '移动版 RTX 4070 使用 AD106 Die，8GB GDDR6，满血功耗 140W。',
    dataQuality: 'official',
    sources: [TPU],
  },
  {
    id: 'nvidia-geforce-rtx-4060-laptop',
    brand: 'nvidia', category: 'gpu', formFactor: 'mobile',
    model: 'GeForce RTX 4060 Laptop',
    codename: 'AD107 (Ada Lovelace)',
    generation: 'RTX 40 Laptop',
    process: 'TSMC 4N (5nm 级)',
    release: '2023-02',
    package: { ...GPU_PKG_MOBILE },
    dies: [{ name: 'AD107 Monolithic Die', role: 'graphics', process: 'TSMC 4N', areaMm2: 158.7, lengthMm: null, widthMm: null, transistorsMillions: 18900, note: '与桌面 RTX 4060 同款 AD107 Die' }],
    transistorsMillions: 18900,
    tdp: 100,
    loadTempRange: '65-78°C',
    notes: '移动版 RTX 4060 使用 AD107 Die，8GB GDDR6，满血功耗 140W。',
    dataQuality: 'official',
    sources: [TPU],
  },
  {
    id: 'nvidia-geforce-rtx-4050-laptop',
    brand: 'nvidia', category: 'gpu', formFactor: 'mobile',
    model: 'GeForce RTX 4050 Laptop',
    codename: 'AD107 (Ada Lovelace)',
    generation: 'RTX 40 Laptop',
    process: 'TSMC 4N (5nm 级)',
    release: '2023-02',
    package: { ...GPU_PKG_MOBILE },
    dies: [{ name: 'AD107 Monolithic Die', role: 'graphics', process: 'TSMC 4N', areaMm2: 158.7, lengthMm: null, widthMm: null, transistorsMillions: 18900, note: '与桌面 RTX 4050 同款 AD107 Die' }],
    transistorsMillions: 18900,
    tdp: 100,
    loadTempRange: '65-78°C',
    notes: '移动版 RTX 4050 使用 AD107 Die，6GB GDDR6，满血功耗 115W。',
    dataQuality: 'official',
    sources: [TPU],
  },
  // --- RTX 30 Laptop (Ampere) ---
  {
    id: 'nvidia-geforce-rtx-3080-ti-laptop',
    brand: 'nvidia', category: 'gpu', formFactor: 'mobile',
    model: 'GeForce RTX 3080 Ti Laptop',
    codename: 'GA103 (Ampere)',
    generation: 'RTX 30 Laptop',
    process: 'Samsung 8N (8nm)',
    release: '2022-01',
    package: { ...GPU_PKG_MOBILE },
    dies: [{ name: 'GA103 Monolithic Die', role: 'graphics', process: 'Samsung 8N', areaMm2: 496, lengthMm: null, widthMm: null, transistorsMillions: 28000, note: '与桌面 RTX 3080 Ti 同款 GA103 Die' }],
    transistorsMillions: 28000,
    tdp: 165,
    loadTempRange: '72-86°C',
    notes: 'RTX 30 系移动旗舰，16GB GDDR6，满血功耗 175W。',
    dataQuality: 'official',
    sources: [TPU],
  },
  {
    id: 'nvidia-geforce-rtx-3080-laptop',
    brand: 'nvidia', category: 'gpu', formFactor: 'mobile',
    model: 'GeForce RTX 3080 Laptop',
    codename: 'GA104 (Ampere)',
    generation: 'RTX 30 Laptop',
    process: 'Samsung 8N (8nm)',
    release: '2021-01',
    package: { ...GPU_PKG_MOBILE },
    dies: [{ name: 'GA104 Monolithic Die', role: 'graphics', process: 'Samsung 8N', areaMm2: 392, lengthMm: null, widthMm: null, transistorsMillions: 17400, note: '与桌面 RTX 3070 同款 GA104 Die' }],
    transistorsMillions: 17400,
    tdp: 150,
    loadTempRange: '70-85°C',
    notes: 'RTX 3080 Laptop，8GB GDDR6，满血功耗 165W。',
    dataQuality: 'official',
    sources: [TPU],
  },
  {
    id: 'nvidia-geforce-rtx-3070-ti-laptop',
    brand: 'nvidia', category: 'gpu', formFactor: 'mobile',
    model: 'GeForce RTX 3070 Ti Laptop',
    codename: 'GA104 (Ampere)',
    generation: 'RTX 30 Laptop',
    process: 'Samsung 8N (8nm)',
    release: '2022-01',
    package: { ...GPU_PKG_MOBILE },
    dies: [{ name: 'GA104 Monolithic Die', role: 'graphics', process: 'Samsung 8N', areaMm2: 392, lengthMm: null, widthMm: null, transistorsMillions: 17400 }],
    transistorsMillions: 17400,
    tdp: 125,
    loadTempRange: '70-85°C',
    notes: 'RTX 3070 Ti Laptop，8GB GDDR6，满血功耗 150W。',
    dataQuality: 'official',
    sources: [TPU],
  },
  {
    id: 'nvidia-geforce-rtx-3070-laptop',
    brand: 'nvidia', category: 'gpu', formFactor: 'mobile',
    model: 'GeForce RTX 3070 Laptop',
    codename: 'GA104 (Ampere)',
    generation: 'RTX 30 Laptop',
    process: 'Samsung 8N (8nm)',
    release: '2021-01',
    package: { ...GPU_PKG_MOBILE },
    dies: [{ name: 'GA104 Monolithic Die', role: 'graphics', process: 'Samsung 8N', areaMm2: 392, lengthMm: null, widthMm: null, transistorsMillions: 17400 }],
    transistorsMillions: 17400,
    tdp: 115,
    loadTempRange: '68-84°C',
    notes: 'RTX 3070 Laptop，8GB GDDR6，满血功耗 140W。',
    dataQuality: 'official',
    sources: [TPU],
  },
  {
    id: 'nvidia-geforce-rtx-3060-laptop',
    brand: 'nvidia', category: 'gpu', formFactor: 'mobile',
    model: 'GeForce RTX 3060 Laptop',
    codename: 'GA106 (Ampere)',
    generation: 'RTX 30 Laptop',
    process: 'Samsung 8N (8nm)',
    release: '2021-01',
    package: { ...GPU_PKG_MOBILE },
    dies: [{ name: 'GA106 Monolithic Die', role: 'graphics', process: 'Samsung 8N', areaMm2: 275.9, lengthMm: null, widthMm: null, transistorsMillions: 12000, note: '与桌面 RTX 3060 同款 GA106 Die' }],
    transistorsMillions: 12000,
    tdp: 115,
    loadTempRange: '70-85°C',
    notes: '移动版 RTX 3060 使用 GA106 Die，6GB GDDR6，满血功耗 130W。',
    dataQuality: 'official',
    sources: [TPU],
  },
  {
    id: 'nvidia-geforce-rtx-3050-ti-laptop',
    brand: 'nvidia', category: 'gpu', formFactor: 'mobile',
    model: 'GeForce RTX 3050 Ti Laptop',
    codename: 'GA107 (Ampere)',
    generation: 'RTX 30 Laptop',
    process: 'Samsung 8N (8nm)',
    release: '2021-05',
    package: { ...GPU_PKG_MOBILE },
    dies: [{ name: 'GA107 Monolithic Die', role: 'graphics', process: 'Samsung 8N', areaMm2: 159, lengthMm: null, widthMm: null, transistorsMillions: 8700 }],
    transistorsMillions: 8700,
    tdp: 75,
    loadTempRange: '65-82°C',
    notes: 'RTX 3050 Ti Laptop，4GB GDDR6，满血功耗 95W。',
    dataQuality: 'official',
    sources: [TPU],
  },
  {
    id: 'nvidia-geforce-rtx-3050-laptop',
    brand: 'nvidia', category: 'gpu', formFactor: 'mobile',
    model: 'GeForce RTX 3050 Laptop',
    codename: 'GA107 (Ampere)',
    generation: 'RTX 30 Laptop',
    process: 'Samsung 8N (8nm)',
    release: '2021-05',
    package: { ...GPU_PKG_MOBILE },
    dies: [{ name: 'GA107 Monolithic Die', role: 'graphics', process: 'Samsung 8N', areaMm2: 159, lengthMm: null, widthMm: null, transistorsMillions: 8700 }],
    transistorsMillions: 8700,
    tdp: 60,
    loadTempRange: '65-82°C',
    notes: 'RTX 3050 Laptop，4GB GDDR6，满血功耗 95W。',
    dataQuality: 'official',
    sources: [TPU],
  },
  // --- GTX 16 系 ---
  {
    id: 'nvidia-geforce-gtx-1660-ti-laptop',
    brand: 'nvidia', category: 'gpu', formFactor: 'mobile',
    model: 'GeForce GTX 1660 Ti',
    codename: 'TU116 (Turing)',
    generation: 'GTX 16 Mobile',
    process: 'TSMC 12nm',
    release: '2019-04',
    package: { ...GPU_PKG_MOBILE },
    dies: [{ name: 'TU116 Monolithic Die', role: 'graphics', process: 'TSMC 12nm', areaMm2: 284, lengthMm: null, widthMm: null, transistorsMillions: 6600 }],
    transistorsMillions: 6600,
    tdp: 80,
    loadTempRange: '65-80°C',
    notes: 'GTX 1660 Ti Laptop，6GB GDDR6。',
    dataQuality: 'official',
    sources: [TPU],
  },
  {
    id: 'nvidia-geforce-gtx-1650-ti-laptop',
    brand: 'nvidia', category: 'gpu', formFactor: 'mobile',
    model: 'GeForce GTX 1650 Ti',
    codename: 'TU117 (Turing)',
    generation: 'GTX 16 Mobile',
    process: 'TSMC 12nm',
    release: '2020-04',
    package: { ...GPU_PKG_MOBILE },
    dies: [{ name: 'TU117 Monolithic Die', role: 'graphics', process: 'TSMC 12nm', areaMm2: 200, lengthMm: null, widthMm: null, transistorsMillions: 4700 }],
    transistorsMillions: 4700,
    tdp: 50,
    loadTempRange: '60-78°C',
    notes: 'GTX 1650 Ti Laptop，4GB GDDR6。',
    dataQuality: 'official',
    sources: [TPU],
  },
  {
    id: 'nvidia-geforce-gtx-1650-laptop',
    brand: 'nvidia', category: 'gpu', formFactor: 'mobile',
    model: 'GeForce GTX 1650',
    codename: 'TU117 (Turing)',
    generation: 'GTX 16 Mobile',
    process: 'TSMC 12nm',
    release: '2019-04',
    package: { ...GPU_PKG_MOBILE },
    dies: [{ name: 'TU117 Monolithic Die', role: 'graphics', process: 'TSMC 12nm', areaMm2: 200, lengthMm: null, widthMm: null, transistorsMillions: 4700 }],
    transistorsMillions: 4700,
    tdp: 50,
    loadTempRange: '60-78°C',
    notes: 'GTX 1650 Laptop，4GB GDDR6。',
    dataQuality: 'official',
    sources: [TPU],
  },
  // --- RTX 20 系 ---
  {
    id: 'nvidia-geforce-rtx-2070-laptop',
    brand: 'nvidia', category: 'gpu', formFactor: 'mobile',
    model: 'GeForce RTX 2070',
    codename: 'TU106 (Turing)',
    generation: 'RTX 20 Mobile',
    process: 'TSMC 12nm',
    release: '2019-04',
    package: { ...GPU_PKG_MOBILE },
    dies: [{ name: 'TU106 Monolithic Die', role: 'graphics', process: 'TSMC 12nm', areaMm2: 445, lengthMm: null, widthMm: null, transistorsMillions: 10800 }],
    transistorsMillions: 10800,
    tdp: 115,
    loadTempRange: '68-84°C',
    notes: 'RTX 2070 Laptop，8GB GDDR6。',
    dataQuality: 'official',
    sources: [TPU],
  },
  {
    id: 'nvidia-geforce-rtx-2060-laptop',
    brand: 'nvidia', category: 'gpu', formFactor: 'mobile',
    model: 'GeForce RTX 2060',
    codename: 'TU106 (Turing)',
    generation: 'RTX 20 Mobile',
    process: 'TSMC 12nm',
    release: '2019-04',
    package: { ...GPU_PKG_MOBILE },
    dies: [{ name: 'TU106 Monolithic Die', role: 'graphics', process: 'TSMC 12nm', areaMm2: 445, lengthMm: null, widthMm: null, transistorsMillions: 10800 }],
    transistorsMillions: 10800,
    tdp: 90,
    loadTempRange: '65-82°C',
    notes: 'RTX 2060 Laptop，6GB GDDR6。',
    dataQuality: 'official',
    sources: [TPU],
  },
  // --- AMD Radeon Mobile ---
  {
    id: 'amd-radeon-rx-6800m',
    brand: 'amd', category: 'gpu', formFactor: 'mobile',
    model: 'Radeon RX 6800M',
    codename: 'Navi 22 (RDNA2)',
    generation: 'RX 6000 Mobile',
    process: 'TSMC 7nm',
    release: '2021-06',
    package: { ...GPU_PKG_MOBILE },
    dies: [{ name: 'Navi 22 Monolithic Die', role: 'graphics', process: 'TSMC 7nm', areaMm2: 335, lengthMm: null, widthMm: null, transistorsMillions: 17200 }],
    transistorsMillions: 17200,
    tdp: 145,
    loadTempRange: '70-85°C',
    notes: 'Radeon RX 6800M，12GB GDDR6，满血功耗 145W。',
    dataQuality: 'official',
    sources: [TPU],
  },
  {
    id: 'amd-radeon-rx-6800s',
    brand: 'amd', category: 'gpu', formFactor: 'mobile',
    model: 'Radeon RX 6800S',
    codename: 'Navi 22 (RDNA2)',
    generation: 'RX 6000 Mobile',
    process: 'TSMC 7nm',
    release: '2022-01',
    package: { ...GPU_PKG_MOBILE },
    dies: [{ name: 'Navi 22 Monolithic Die', role: 'graphics', process: 'TSMC 7nm', areaMm2: 335, lengthMm: null, widthMm: null, transistorsMillions: 17200 }],
    transistorsMillions: 17200,
    tdp: 100,
    loadTempRange: '68-82°C',
    notes: 'Radeon RX 6800S（低功耗版），8GB GDDR6。',
    dataQuality: 'official',
    sources: [TPU],
  },
  {
    id: 'amd-radeon-rx-6700s',
    brand: 'amd', category: 'gpu', formFactor: 'mobile',
    model: 'Radeon RX 6700S',
    codename: 'Navi 22 (RDNA2)',
    generation: 'RX 6000 Mobile',
    process: 'TSMC 7nm',
    release: '2022-01',
    package: { ...GPU_PKG_MOBILE },
    dies: [{ name: 'Navi 22 Monolithic Die', role: 'graphics', process: 'TSMC 7nm', areaMm2: 335, lengthMm: null, widthMm: null, transistorsMillions: 17200 }],
    transistorsMillions: 17200,
    tdp: 100,
    loadTempRange: '68-82°C',
    notes: 'Radeon RX 6700S（低功耗版），10GB GDDR6。',
    dataQuality: 'official',
    sources: [TPU],
  },
  {
    id: 'amd-radeon-rx-7600m-xt',
    brand: 'amd', category: 'gpu', formFactor: 'mobile',
    model: 'Radeon RX 7600M XT',
    codename: 'Navi 33 (RDNA3)',
    generation: 'RX 7000 Mobile',
    process: 'TSMC 6nm',
    release: '2023-01',
    package: { ...GPU_PKG_MOBILE },
    dies: [{ name: 'Navi 33 Monolithic Die', role: 'graphics', process: 'TSMC 6nm', areaMm2: 204, lengthMm: null, widthMm: null, transistorsMillions: 13300 }],
    transistorsMillions: 13300,
    tdp: 120,
    loadTempRange: '70-85°C',
    notes: '移动版 RX 7600M XT，8GB GDDR6，RDNA3 架构，满血功耗 120W。',
    dataQuality: 'official',
    sources: [TPU],
  },
'''


# ============================================================
# 2. Generate laptops.ts from 游戏本.xlsx
# ============================================================

BRAND_MAP = {
    '联想': 'lenovo',
    '华硕': 'asus',
    '惠普': 'hp',
    '戴尔': 'dell',
    '宏碁': 'acer',
    '微星': 'msi',
    '雷蛇': 'razer',
    '七彩虹': 'colorful',
    '机械革命': 'mechrevo',
    '神舟': 'hasee',
    '小米': 'xiaomi',
    '荣耀': 'honor',
    '技嘉': 'gigabyte',
    '华为': 'huawei',
    '机械师': 'machenike',
    '雷神': 'thunderobot',
}

BRAND_LABELS = {
    'lenovo': '联想 Lenovo',
    'asus': '华硕 ASUS',
    'hp': '惠普 HP',
    'dell': '戴尔 Dell',
    'acer': '宏碁 Acer',
    'msi': '微星 MSI',
    'razer': '雷蛇 Razer',
    'colorful': '七彩虹 Colorful',
    'mechrevo': '机械革命 Mechrevo',
    'hasee': '神舟 HASEE',
    'xiaomi': '小米 Xiaomi',
    'honor': '荣耀 Honor',
    'gigabyte': '技嘉 GIGABYTE',
    'huawei': '华为 Huawei',
    'machenike': '机械师 MACHENIKE',
    'thunderobot': '雷神 Thunderobot',
}

def cpu_platform(cpu_str):
    """Determine if CPU is Intel or AMD."""
    s = str(cpu_str).lower()
    if 'r9' in s or 'r7' in s or 'r5' in s or 'ryzen' in s or '锐龙' in s or 'amd' in s:
        return 'AMD'
    return 'Intel'

def gen_laptops():
    wb = openpyxl.load_workbook('D:/Desktop/游戏本.xlsx', data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[0] is None:
            continue
        rows.append(dict(zip(headers, row)))

    # 中文名称规范化：去空格/去年份/去平台与变体后缀，用于合并同型号不同配置
    def norm_name(n):
        s = str(n).replace(' ', '')
        s = re.sub(r'\d{4}款?$', '', s)   # 去末尾年份（2021 / 2021款）
        s = s.replace('AI元启', '')
        s = s.replace('酷睿版', '').replace('锐龙版', '')
        return s

    # Group by (brand, 规范化中文名称, 型号)
    from collections import defaultdict, OrderedDict
    groups = OrderedDict()

    for r in rows:
        brand_cn = str(r.get('品牌', '')).strip()
        series = str(r.get('系列', '')).strip()
        name = norm_name(r.get('中文名称', ''))
        model = str(r.get('型号', '')).strip()
        key = (brand_cn, name, model)

        if key not in groups:
            groups[key] = {
                'brand_cn': brand_cn,
                'series': series,
                'name': name,
                'model': model,
                'release': None,
                'cpus': [],
                'gpus': [],
                'rams': [],
                'storages': [],
                'displays': [],
            }

        g = groups[key]
        release = str(r.get('发布时间', '')).strip()
        # 取组内最早发布时间，保证年份为系列起始年份
        if release and (not g['release'] or release < g['release']):
            g['release'] = release

        cpu = str(r.get('处理器', '')).strip()
        gpu = str(r.get('显卡', '')).strip()
        ram = str(r.get('内存', '')).strip()
        storage = str(r.get('硬盘', '')).strip()
        display = str(r.get('屏幕', '')).strip()

        if cpu and cpu not in g['cpus']:
            g['cpus'].append(cpu)
        if gpu and gpu not in g['gpus']:
            g['gpus'].append(gpu)
        if ram and ram not in g['rams']:
            g['rams'].append(ram)
        if storage and storage not in g['storages']:
            g['storages'].append(storage)
        if display and display not in g['displays']:
            g['displays'].append(display)

    # Generate TypeScript
    lines = []
    lines.append("import type { Laptop } from './types';\n")
    lines.append("// Auto-generated from 游戏本.xlsx — 游戏本配置数据 (2020-2026)")
    lines.append("// 数据来源：品牌官网 / 京东 / 天猫等公开渠道\n")
    lines.append("export const allLaptops: Laptop[] = [")

    used_ids = set()

    def unique_id(base):
        """保证 id 全局唯一：冲突时追加 -2 / -3 ..."""
        i = 2
        cand = base
        while cand in used_ids:
            cand = f"{base}-{i}"
            i += 1
        used_ids.add(cand)
        return cand

    for key, g in groups.items():
        brand = BRAND_MAP.get(g['brand_cn'], 'other')
        name = g['name']
        model = g['model']
        series = g['series']

        # Generate ID (保证唯一)
        id_str = unique_id(f"{brand}-{slugify(name)}-{slugify(model)}")

        # Representative values (first non-empty)
        ram = g['rams'][0] if g['rams'] else '—'
        storage = g['storages'][0] if g['storages'] else '—'
        display = g['displays'][0] if g['displays'] else '—'
        release = g['release'] if g['release'] else None

        lines.append("  {")
        lines.append(f"    id: '{id_str}',")
        lines.append(f"    brand: '{brand}',")
        lines.append(f"    series: '{series}',")
        lines.append(f"    displayName: '{name}',")
        lines.append(f"    model: '{model}',")
        release_ts = f"'{release}'" if release else 'null'
        lines.append(f"    release: {release_ts},")

        # CPU options
        cpu_items = [f"'{c}'" for c in g['cpus']]
        lines.append(f"    cpuOptions: [{', '.join(cpu_items)}],")

        # GPU options
        gpu_items = [f"'{gpu}'" for gpu in g['gpus']]
        lines.append(f"    gpuOptions: [{', '.join(gpu_items)}],")

        lines.append(f"    ram: '{ram}',")
        lines.append(f"    storage: '{storage}',")
        lines.append(f"    display: '{display}',")
        lines.append(f"    weightKg: null,")
        lines.append(f"    sources: [],")
        lines.append("  },")

    lines.append("];\n")
    lines.append("export function getLaptopById(id: string): Laptop | undefined {")
    lines.append("  return allLaptops.find((l) => l.id === id);")
    lines.append("}")
    return '\n'.join(lines)


# ============================================================
# Run
# ============================================================

if __name__ == '__main__':
    chips_ts = gen_mobile_chips()
    with open(os.path.join(OUT_DIR, '..', 'src', 'data', 'chips', 'mobile-chips.ts'), 'w', encoding='utf-8') as f:
        f.write(chips_ts)
    print(f"Generated mobile-chips.ts")

    laptops_ts = gen_laptops()
    with open(os.path.join(OUT_DIR, '..', 'src', 'data', 'laptops.ts'), 'w', encoding='utf-8') as f:
        f.write(laptops_ts)
    print(f"Generated laptops.ts")
